"""
Excel(セグロット在庫表)から日次の入出庫を一括取り込む。

バーコード入力が現場の負荷として重すぎるとのフィードバックにより、セグロットには
従来通りExcelで入出庫を記録してもらい、それをアップロードして取り込む運用に戻すための機能。

想定シート構造(【フェリクロス】セグロット在庫表.xls 等の月別タブ):
  行0: 商品コード, 商品名, "", 入数, 前月残, <日付シリアル>, "", "", <日付シリアル>, "", "", ...
  行1: "", "", JAN, "", "", 入荷, 出荷, 在庫, 入荷, 出荷, 在庫, ...
  行2以降: 各商品の行。3列1組で日付ごとの[入荷,出荷,在庫]が入る。

指定日以前の全ブロックについて、各商品の入荷・出荷を読み取り、(商品コード, 日付)単位の
movement一覧として返す。取り込み側(db.replace_excel_movements)で、同じ(商品コード,日付)の
既存の取り込み分は置き換えるため、同じシートを再アップロードして修正しても安全。
"""
import io
from datetime import datetime, timedelta

EXCEL_EPOCH = datetime(1899, 12, 30)


def _serial_to_date(serial):
    try:
        return (EXCEL_EPOCH + timedelta(days=float(serial))).date()
    except (TypeError, ValueError):
        return None


def _load_sheet_rows(file_bytes, filename, sheet_name):
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext == "xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=file_bytes)
        if sheet_name not in book.sheet_names():
            raise ValueError(f"シート「{sheet_name}」が見つかりません（存在するシート: {', '.join(book.sheet_names())}）")
        sheet = book.sheet_by_name(sheet_name)
        return [sheet.row_values(r) for r in range(sheet.nrows)]
    elif ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート「{sheet_name}」が見つかりません（存在するシート: {', '.join(wb.sheetnames)}）")
        ws = wb[sheet_name]
        return [[c.value for c in row] for row in ws.iter_rows()]
    else:
        raise ValueError("xlsまたはxlsxファイルをアップロードしてください")


def parse_daily_movements(file_bytes, filename, sheet_name, up_to_date):
    """
    up_to_date: date オブジェクト。シート内の最初の日付〜up_to_dateの入荷・出荷を集計する。
    戻り値: {"movements": [{"code":.., "date":"YYYY-MM-DD", "kind":"in"/"out", "qty":int}, ...],
             "date_from": "YYYY-MM-DD", "date_to": "YYYY-MM-DD"}
    """
    rows = _load_sheet_rows(file_bytes, filename, sheet_name)
    if len(rows) < 3:
        raise ValueError("シートの形式が想定と異なります（データ行が見つかりません）")

    header = rows[0]
    blocks = []  # (date, in_col, out_col)
    col = 5
    while col + 2 < len(header):
        d = _serial_to_date(header[col])
        if d is not None and d <= up_to_date:
            blocks.append((d, col, col + 1))
        col += 3
    if not blocks:
        raise ValueError("指定日以前の日付データがシート内に見つかりません")

    day_totals = {}  # code -> {"YYYY-MM-DD": {"in":qty, "out":qty}}
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        if not code:
            continue
        day_map = day_totals.setdefault(code, {})
        for d, in_col, out_col in blocks:
            in_val = row[in_col] if in_col < len(row) else None
            out_val = row[out_col] if out_col < len(row) else None
            in_qty = int(in_val) if isinstance(in_val, (int, float)) and in_val else 0
            out_qty = int(out_val) if isinstance(out_val, (int, float)) and out_val else 0
            if in_qty == 0 and out_qty == 0:
                continue
            key = d.isoformat()
            cur = day_map.setdefault(key, {"in": 0, "out": 0})
            cur["in"] += in_qty
            cur["out"] += out_qty

    movements = []
    for code, day_map in day_totals.items():
        for dt, vals in day_map.items():
            if vals["in"] > 0:
                movements.append({"code": code, "date": dt, "kind": "in", "qty": vals["in"]})
            if vals["out"] > 0:
                movements.append({"code": code, "date": dt, "kind": "out", "qty": vals["out"]})

    return {
        "movements": movements,
        "date_from": blocks[0][0].isoformat(),
        "date_to": blocks[-1][0].isoformat(),
    }
